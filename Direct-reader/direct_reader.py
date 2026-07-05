#!/usr/bin/env python3
"""
Progressive Direct Interview Reader
Extracts field inventory from PAA and HQX Direct Interview Excel files
into a unified master CSV (Layer 1 of the migration pipeline).

Usage:
    python3 direct_reader.py

Input files expected at /home/claude/ (copy from /mnt/project/ first):
    PAA_2_0_Direct_Interview.xlsx
    HQX2_0_Direct_Interview.xlsx

Output:
    /home/claude/Progressive_Direct_Master.csv
    /home/claude/Progressive_Direct_Warnings.csv  (if any issues found)

Improvements applied:
    - Section: and Disclaimer: rows explicitly skipped
    - LOB_MISSING flagged and written to Warnings CSV
    - Page names normalized (double spaces collapsed)
    - Version column captured for direct-sync use
"""

import re
import pandas as pd
from openpyxl import load_workbook

# ── Config ────────────────────────────────────────────────────────────────────

FILES = [
    {
        'path': '/home/claude/PAA_2_0_Direct_Interview.xlsx',
        'flow': 'Agent',
        'source': 'PAA',
        'skip_sheets': {'DI Guide', 'Change Control'},
        'carrier_sheet': 'Carrier Questions',
    },
    {
        'path': '/home/claude/HQX2_0_Direct_Interview.xlsx',
        'flow': 'Consumer',
        'source': 'HQX',
        'skip_sheets': {'Version Control', 'Three Prefill Questions'},
        'carrier_sheet': None,
    },
]

OUTPUT_PATH    = '/home/claude/Progressive_Direct_Master.csv'
WARNINGS_PATH  = '/home/claude/Progressive_Direct_Warnings.csv'

LOB_COLUMNS = {'HO3', 'DF', 'MFH', 'Condo'}
LOB_MAP     = {'HO3': 'HO3', 'DF': 'DF', 'MFH': 'MFH', 'Condo': 'HO6'}

CARRIER_COLUMNS = {
    'HSI', 'Hippo', 'Bamboo', 'PGRH/ASI', 'Foremost', 'AMod',
    'NTW', 'NatGen', 'Assurant', 'Openly', 'P Rock',
    'American Integrity', 'Stillwater', 'Tower Hill'
}

COLUMN_MAP = {
    'getquote data dictionary':          'GetQuote_DD',
    'field id':                          'Field ID',
    'pre-fill element':                  'Pre-fill Element',
    'label':                             'Label',
    ' label':                            'Label',
    'control type':                      'Control Type',
    ' control type':                     'Control Type',
    'watermarks':                        'Watermarks',
    'learn more':                        'Learn More',
    'quote flow - new / revised':        'Quote Flow',
    'display rules':                     'Display Rules',
    'verification indicator displayed?': 'Verification Indicator',
    'verification indicator displayed':  'Verification Indicator',
    'default value':                     'Default Value',
    'consumer default value':            'Default Value',
    'parent or child question':          'Parent or Child',
    'presented if vendor prefilled':     'Presented if Prefilled',
    'presented if vendor not prefilled': 'Presented if Not Prefilled',
    'mandatory - if question displays':  'Mandatory',
    'field format':                      'Field Format',
    'value list':                        'Value List',
    'bolt error message':                'BOLT Error Message',
    'kickout':                           'Kickout',
    'prefill values':                    'Prefill Values',
    'max characters from pre-fill':      'Max Prefill Chars',
    'legacy mapping':                    'Legacy Mapping',
    '2.0 consumer mapping':              'Consumer Mapping',
    'version':                           'Version',
    'value for rc1 submission ':         'RC1 Value',
}

OUTPUT_COLUMNS = [
    'Field ID', 'Canonical Name', 'Pre-fill Element', 'Page', 'Flow', 'Source File',
    'LOBs', 'Carrier', 'Label', 'Control Type', 'Value List', 'Field Format',
    'Default Value', 'Display Rules', 'Mandatory', 'BOLT Error Message',
    'Kickout', 'Parent or Child', 'Prefill Values',
    'Presented if Prefilled', 'Presented if Not Prefilled',
    'Quote Flow', 'Version', 'Legacy Mapping', 'Consumer Mapping',
    'GetQuote_DD',
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize_page(name):
    """Normalize page name — trim and collapse double spaces."""
    return re.sub(r'  +', ' ', str(name).strip())


def canonical_name(field_id, getquote_dd):
    if field_id and str(field_id).strip() not in ('N/A', '', 'None'):
        fid = str(field_id).strip()
        m = re.match(r'^PL_F\d+_(.+)$', fid)
        if m:
            return m.group(1)
        m = re.match(r'^PL_(.+)$', fid)
        if m:
            return m.group(1)
        return fid
    if getquote_dd and str(getquote_dd).strip() not in ('N/A', '', 'None'):
        return str(getquote_dd).strip()
    return ''


def normalize_headers(headers):
    return {str(h).strip().lower(): str(h).strip() for h in headers if h is not None}


def get_lobs(row, header_map, field_index=None):
    """
    Extract LOB checkmarks. field_index: if set, row was split from a multi-field
    cell — pick the nth value from newline-separated LOB cells (e.g. 'v\nv\nv').
    """
    lobs = []
    for col_lower, col_orig in header_map.items():
        stripped = col_lower.strip()
        if stripped in {k.lower() for k in LOB_COLUMNS}:
            key = next((k for k in LOB_COLUMNS if k.lower() == stripped), None)
            if key:
                val = str(row.get(col_orig, '') or '')
                if field_index is not None:
                    parts = [p.strip().lower() for p in val.split('\n')]
                    if len(parts) == 1:
                        # Single v covers all fields in the group
                        marked = parts[0] == 'v'
                    else:
                        # One v per field — pick by index
                        marked = field_index < len(parts) and parts[field_index] == 'v'
                else:
                    marked = val.strip().lower() == 'v'
                if marked:
                    lobs.append(LOB_MAP[key])
    if set(lobs) == set(LOB_MAP.values()):
        return 'All'
    return ', '.join(lobs) if lobs else 'LOB_MISSING'


def get_carriers(row, header_map):
    carriers = []
    for col_lower, col_orig in header_map.items():
        if col_orig.strip() in CARRIER_COLUMNS:
            val = row.get(col_orig, '')
            if str(val).strip().lower() == 'v':
                carriers.append(col_orig.strip())
    return ', '.join(carriers) if carriers else ''


def is_skip_row(field_id, label, getquote_dd=None):
    fid = str(field_id).strip() if field_id else ''
    lbl = str(label).strip() if label else ''
    gq  = str(getquote_dd).strip() if getquote_dd else ''
    # Section and Disclaimer rows — Field ID column
    if fid.startswith('Section:') or fid.startswith('Disclaimer:'):
        return True
    # Section rows where the label is in GetQuote DD column instead (PAA Overview pattern)
    if gq.startswith('Section:') or gq in ('Selected Carrier', 'Prefill Verification', 'Carrier Questions'):
        return True
    # UI text blocks with no Field ID
    if fid in ('N/A', '', 'None') and any(phrase in lbl for phrase in [
        'We have everything we need',
        'Click continue',
        'A few key points',
        'key points',
    ]):
        return True
    if fid in ('N/A', '', 'None') and not lbl:
        return True
    return False


def split_multi_field_row(row, header_map=None, raw_row=None):
    field_id = str(row.get('Field ID', '')).strip()
    prefill  = str(row.get('Pre-fill Element', '')).strip()
    getquote = str(row.get('GetQuote_DD', '')).strip()

    field_ids = [f.strip() for f in field_id.split('\n') if f.strip() and f.strip() not in ('N/A', '')]
    prefills  = [p.strip() for p in prefill.split('\n') if p.strip() and p.strip() not in ('N/A', '')]
    getquotes = [g.strip() for g in re.split(r'\n', getquote)   if g.strip() and g.strip() not in ('N/A', '')]

    if len(field_ids) <= 1:
        return [row]

    rows = []
    seen_ids = set()
    for i, fid in enumerate(field_ids):
        if fid in seen_ids:
            continue
        seen_ids.add(fid)
        new_row = dict(row)
        new_row['Field ID']         = fid
        new_row['Pre-fill Element'] = prefills[i]  if i < len(prefills)  else ''
        new_row['GetQuote_DD']      = getquotes[i] if i < len(getquotes) else ''
        new_row['Canonical Name']   = canonical_name(fid, new_row['GetQuote_DD'])
        # Re-derive LOBs using the original Excel row so LOB columns are found correctly
        if header_map is not None and raw_row is not None:
            new_row['LOBs'] = get_lobs(raw_row, header_map, field_index=i)
        rows.append(new_row)
    return rows


# ── Sheet Reader ──────────────────────────────────────────────────────────────

def read_sheet(ws, page, flow, source, is_carrier=False):
    rows_out    = []
    skipped     = 0
    split_count = 0

    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return rows_out, skipped, split_count

    # Find header row
    header_row_idx = None
    for i, row in enumerate(raw_rows):
        cells = [str(c).strip().lower() for c in row if c is not None]
        if 'field id' in cells:
            header_row_idx = i
            break

    if header_row_idx is None:
        return rows_out, skipped, split_count

    headers      = raw_rows[header_row_idx]
    header_map   = normalize_headers(headers)
    orig_headers = [str(h).strip() if h else '' for h in headers]

    for raw_row in raw_rows[header_row_idx + 1:]:
        if all(c is None or str(c).strip() == '' for c in raw_row):
            continue

        row_dict = {}
        for i, val in enumerate(raw_row):
            if i < len(orig_headers) and orig_headers[i]:
                row_dict[orig_headers[i]] = val

        norm = {}
        for orig_key, val in row_dict.items():
            mapped = COLUMN_MAP.get(orig_key.lower(), orig_key)
            if mapped not in norm:
                norm[mapped] = val

        field_id = str(norm.get('Field ID', '')).strip()
        label    = str(norm.get('Label', '')).strip()

        if is_skip_row(field_id, label, norm.get('GetQuote_DD', '')):
            skipped += 1
            continue

        norm['Page']         = page  # already normalized before passing in
        norm['Flow']         = flow
        norm['Source File']  = source
        norm['LOBs']         = get_lobs(row_dict, header_map)
        norm['Carrier']      = get_carriers(row_dict, header_map) if is_carrier else ''
        norm['Canonical Name'] = canonical_name(field_id, norm.get('GetQuote_DD', ''))

        split_rows = split_multi_field_row(norm, header_map=header_map, raw_row=row_dict)
        if len(split_rows) > 1:
            split_count += 1
        rows_out.extend(split_rows)

    return rows_out, skipped, split_count


# ── Main ──────────────────────────────────────────────────────────────────────

def run():
    all_rows = []
    stats    = []

    for file_cfg in FILES:
        path         = file_cfg['path']
        flow         = file_cfg['flow']
        source       = file_cfg['source']
        skip_sheets  = file_cfg['skip_sheets']
        carrier_sheet = file_cfg['carrier_sheet']

        print(f"\nReading {source} ({flow}): {path}")
        wb = load_workbook(path, read_only=True)

        for sheet_name in wb.sheetnames:
            if sheet_name in skip_sheets:
                print(f"  Skipping: {sheet_name}")
                continue

            # Normalize page name before passing to sheet reader
            page       = normalize_page(sheet_name)
            ws         = wb[sheet_name]
            is_carrier = (sheet_name == carrier_sheet)
            rows, skipped, split = read_sheet(ws, page, flow, source, is_carrier)

            all_rows.extend(rows)
            stats.append({
                'Source':       source,
                'Sheet':        sheet_name,
                'Page':         page,
                'Rows':         len(rows),
                'Skipped':      skipped,
                'Split':        split,
                'Carrier Sheet': is_carrier,
            })
            normalized_note = f" → normalized to '{page}'" if page != sheet_name else ''
            print(f"  {sheet_name}{normalized_note}: {len(rows)} rows, {skipped} skipped, {split} split")

    # Build DataFrame
    df = pd.DataFrame(all_rows)

    for col in OUTPUT_COLUMNS:
        if col not in df.columns:
            df[col] = ''

    df = df[OUTPUT_COLUMNS + [c for c in df.columns if c not in OUTPUT_COLUMNS]]
    df = df.fillna('').astype(str).replace('None', '').replace('nan', '')

    # Write main output
    df.to_csv(OUTPUT_PATH, index=False, encoding='utf-8-sig')

    # Write warnings CSV for LOB_MISSING rows
    warnings = df[df['LOBs'] == 'LOB_MISSING'][['Field ID', 'Canonical Name', 'Page', 'Flow', 'Source File']].copy()
    warnings['Warning'] = 'No LOB checkmarks found — investigate'
    if not warnings.empty:
        warnings.to_csv(WARNINGS_PATH, index=False, encoding='utf-8-sig')

    # Summary
    print(f"\n{'='*60}")
    print(f"Output:        {OUTPUT_PATH}")
    print(f"Total rows:    {len(df)}")
    print(f"Agent rows:    {len(df[df['Flow'] == 'Agent'])}")
    print(f"Consumer rows: {len(df[df['Flow'] == 'Consumer'])}")
    print(f"\nRows per page:")
    for _, s in pd.DataFrame(stats).iterrows():
        print(f"  [{s['Source']}] {s['Page']}: {s['Rows']} rows")

    if not warnings.empty:
        print(f"\nWARNINGS — {len(warnings)} rows with LOB_MISSING → {WARNINGS_PATH}")
        for _, r in warnings.iterrows():
            print(f"  [{r['Source File']}] {r['Page']}: {r['Field ID']}")
    else:
        print(f"\nNo warnings.")


if __name__ == '__main__':
    run()
